"""
SBC B2 BEL Collection Dashboard
================================
Enterprise-grade Streamlit dashboard for collections / field visit / curing
performance monitoring, styled with Security Bank corporate branding.

Run with:
    streamlit run app.py
"""

import io
import re
import unicodedata
from datetime import datetime

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

try:
    from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode
    HAS_AGGRID = True
except Exception:
    HAS_AGGRID = False

# ============================================================================
# PAGE CONFIG & SECURITY BANK THEME
# ============================================================================

st.set_page_config(
    page_title="SBC B2 BEL Collection Dashboard",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ---- Security Bank corporate palette ----
SB_BLUE = "#0057B8"
SB_DARK_BLUE = "#003B7A"
SB_LIGHT_BLUE = "#EAF3FF"
SB_WHITE = "#FFFFFF"
SB_GREEN = "#00A651"
SB_ORANGE = "#F5A623"
SB_RED = "#D32F2F"

STATUS_COLORS = {
    "Active": SB_BLUE,
    "Cured": SB_GREEN,
    "Flowed": SB_RED,
}

BLUES_SCALE = ["#EAF3FF", "#B3D4FF", "#66A3FF", "#0057B8", "#003B7A"]
GREENS_SCALE = ["#E6F7EE", "#99E0BC", "#33C285", "#00A651", "#00753A"]
REDS_SCALE = ["#FCE9E9", "#F0A8A8", "#E36767", "#D32F2F", "#8E1F1F"]

CUSTOM_CSS = f"""
<style>
#MainMenu {{visibility: hidden;}}
footer {{visibility: hidden;}}
footer:after {{content:''; visibility: hidden;}}

.stApp {{ background-color: #F5F8FC; }}

/* ---- Header banner ---- */
.sbc-header {{
    background: linear-gradient(120deg, {SB_DARK_BLUE} 0%, {SB_BLUE} 100%);
    padding: 22px 30px;
    border-radius: 14px;
    margin-bottom: 18px;
    box-shadow: 0 4px 14px rgba(0,59,122,0.25);
    display: flex;
    align-items: center;
    gap: 18px;
}}
.sbc-logo {{
    width: 52px; height: 52px; border-radius: 10px;
    background: {SB_WHITE};
    display: flex; align-items: center; justify-content: center;
    font-size: 26px; font-weight: 800; color: {SB_BLUE};
    box-shadow: 0 2px 6px rgba(0,0,0,0.2);
    flex-shrink: 0;
}}
.sbc-header-text h1 {{
    color: {SB_WHITE} !important; margin: 0; font-size: 26px; font-weight: 800;
    letter-spacing: 0.3px;
}}
.sbc-header-text p {{
    color: {SB_LIGHT_BLUE} !important; margin: 2px 0 0 0; font-size: 13.5px;
}}

/* ---- Sidebar ---- */
section[data-testid="stSidebar"] {{
    background: linear-gradient(180deg, {SB_LIGHT_BLUE} 0%, #CFE4FB 100%);
    border-right: 1px solid #B9D4F2;
}}
section[data-testid="stSidebar"] * {{ color: {SB_DARK_BLUE} !important; }}
section[data-testid="stSidebar"] h1, section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 {{ color: {SB_DARK_BLUE} !important; font-weight: 800 !important; }}
section[data-testid="stSidebar"] .stMultiSelect [data-baseweb="tag"] {{
    background-color: {SB_BLUE} !important;
}}
section[data-testid="stSidebar"] .stMultiSelect [data-baseweb="tag"] span {{
    color: {SB_WHITE} !important;
}}
section[data-testid="stSidebar"] div[data-baseweb="select"] > div {{
    background-color: {SB_WHITE};
    border: 1px solid #B9D4F2;
    border-radius: 8px;
}}
section[data-testid="stSidebar"] button {{
    background-color: {SB_BLUE} !important;
    color: {SB_WHITE} !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    border: none !important;
}}
section[data-testid="stSidebar"] div[data-testid="stFileUploaderDropzone"] {{
    background-color: {SB_WHITE};
    border: 1px dashed #8FB8E8;
    border-radius: 10px;
}}
section[data-testid="stSidebar"] div[data-testid="stExpander"] {{
    background-color: {SB_WHITE};
    border-radius: 10px;
    border: 1px solid #B9D4F2;
}}

/* ---- KPI cards ---- */
div[data-testid="stMetric"] {{
    background: {SB_WHITE};
    border-radius: 12px;
    padding: 16px 18px 12px 18px;
    box-shadow: 0 2px 10px rgba(0,59,122,0.10);
    border-left: 5px solid {SB_BLUE};
    transition: transform 0.15s ease, box-shadow 0.15s ease;
}}
div[data-testid="stMetric"]:hover {{
    transform: translateY(-2px);
    box-shadow: 0 6px 16px rgba(0,59,122,0.18);
}}
div[data-testid="stMetric"] label {{ color: {SB_DARK_BLUE} !important; font-weight: 600 !important; }}
div[data-testid="stMetric"] div[data-testid="stMetricValue"] {{ color: {SB_BLUE} !important; font-weight: 800 !important; }}

/* ---- Headings ---- */
h1, h2, h3 {{ color: {SB_DARK_BLUE}; font-weight: 800; }}
h4, h5 {{ color: {SB_DARK_BLUE}; font-weight: 700; }}

/* ---- Tabs ---- */
.stTabs [data-baseweb="tab-list"] {{
    gap: 4px; background-color: {SB_LIGHT_BLUE}; padding: 6px; border-radius: 12px;
}}
.stTabs [data-baseweb="tab"] {{
    height: 42px; border-radius: 8px; padding: 0 16px; font-weight: 600;
    color: {SB_DARK_BLUE}; background-color: transparent;
}}
.stTabs [data-baseweb="tab"]:hover {{ background-color: rgba(0,87,184,0.12); }}
.stTabs [aria-selected="true"] {{
    background-color: {SB_BLUE} !important; color: {SB_WHITE} !important;
}}

/* ---- Dataframes / tables ---- */
div[data-testid="stDataFrame"] {{
    border-radius: 10px; overflow: hidden; box-shadow: 0 2px 8px rgba(0,59,122,0.08);
}}

/* ---- Containers / expanders ---- */
div[data-testid="stExpander"] {{
    background-color: {SB_WHITE}; border-radius: 10px;
    box-shadow: 0 1px 6px rgba(0,59,122,0.08);
}}
.block-container {{ padding-top: 1.2rem; }}
</style>
"""
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

PHP = lambda v: f"₱{v:,.2f}" if pd.notna(v) else "₱0.00"


def fnum(v):
    if pd.isna(v):
        return "0"
    return f"{v:,.0f}"


def pct(v):
    if pd.isna(v):
        return "0.0%"
    return f"{v:.1f}%"


def traffic_light_color(value, good=70, warn=40, higher_is_better=True):
    """Return a Security Bank brand color (green/amber/red) for a rate value against
    default performance thresholds. Thresholds are reasonable defaults, not audited
    targets, and can be adjusted per metric via the good/warn parameters."""
    if value is None or pd.isna(value):
        return SB_ORANGE
    if higher_is_better:
        if value >= good:
            return SB_GREEN
        if value >= warn:
            return SB_ORANGE
        return SB_RED
    else:
        if value <= good:
            return SB_GREEN
        if value <= warn:
            return SB_ORANGE
        return SB_RED


def render_rate_metric(col, label, value_str, rate_for_color, good=70, warn=40,
                        higher_is_better=True, help_text=None):
    """Render an st.metric plus a thin colored performance bar underneath it
    (green/amber/red) based on traffic_light_color. Use for rate-based KPIs
    where a target/threshold is meaningful; use plain st.metric for raw counts."""
    with col:
        col.metric(label, value_str, help=help_text)
        color = traffic_light_color(rate_for_color, good=good, warn=warn, higher_is_better=higher_is_better)
        st.markdown(
            f'<div style="height:6px;border-radius:3px;background:{color};'
            f'margin-top:-10px;margin-bottom:6px;"></div>',
            unsafe_allow_html=True,
        )


# ============================================================================
# COLUMN RESOLUTION
# ============================================================================

COLUMN_ALIASES = {
    "concat": ["concat", "concat id", "account id", "acct id"],
    "account_name": ["account name", "accountname", "client name", "name"],
    "ob_tad": ["ob/tad", "ob tad", "ob", "tad", "outstanding balance", "balance"],
    "status": ["status"],
    "substatus": ["substatus", "sub status", "sub-status"],
    "area": ["area"],
    "area2": ["area2", "area 2", "area break"],
    "sub_area": ["sub area", "subarea", "province"],
    "industry": ["industry"],
    "agent": ["agent", "collector", "field agent"],
    "action_date": ["action date"],
    "endorsement_date": ["date of endorsement", "endorsement date", "date endorsed"],
    "active": ["active", "active status"],
    "remarks": ["remarks"],
    "payment": ["payment"],
    "openci": ["openci", "open ci"],
    "fv_result": ["fv_result", "fv result", "field visit result"],
    "tele_field": ["tele/field", "tele field", "channel"],
    "area_list": ["area list"],
    "cured_flag": ["cured/not cured", "cured not cured", "cured status"],
    "risk_level": ["risk level", "risk lvl", "bom/fresh", "bom fresh"],
    "bal_distro": ["bal distro", "balance distro", "bal distribution"],
    "total_visits": ["total of visits made", "total visits", "visits made"],
    "ptp_date": ["ptp-date", "ptp date", "ptp_date", "promise to pay date"],
    "ptp_amount": ["ptp-amount", "ptp amount", "ptp_amount", "promise to pay amount"],
    "bp_flag": ["bp?", "bp", "broken promise"],
    "rfd": ["rfd", "reason for default", "reason for disconnection"],
}

REQUIRED_MIN = ["concat", "ob_tad", "status"]


def resolve_columns(columns):
    """Map logical field -> actual dataframe column name (best match)."""
    used = set()
    norm = {c: re.sub(r"\s+", " ", str(c).strip().lower()) for c in columns}
    mapping = {}
    for logical, aliases in COLUMN_ALIASES.items():
        found = None
        for alias in aliases:
            for col, ncol in norm.items():
                if col in used:
                    continue
                if ncol == alias:
                    found = col
                    break
            if found:
                break
        if not found:
            for alias in aliases:
                for col, ncol in norm.items():
                    if col in used:
                        continue
                    if alias in ncol:
                        found = col
                        break
                if found:
                    break
        if found:
            mapping[logical] = found
            used.add(found)
    return mapping


def clean_numeric(series):
    s = series.astype(str).str.replace(r"[₱$,]", "", regex=True).str.strip()
    s = s.replace({"": np.nan, "nan": np.nan, "None": np.nan, "-": np.nan})
    return pd.to_numeric(s, errors="coerce").fillna(0.0)


def clean_date(series):
    return pd.to_datetime(series, errors="coerce")


@st.cache_data(show_spinner=False, max_entries=3)
def load_file(file_bytes, file_name):
    """Read uploaded file bytes into a DataFrame."""
    if file_name.lower().endswith(".csv"):
        df = pd.read_csv(io.BytesIO(file_bytes), low_memory=False)
    else:
        df = pd.read_excel(io.BytesIO(file_bytes))
    df.columns = [str(c).strip() for c in df.columns]
    return df


@st.cache_data(show_spinner=False, max_entries=3)
def prepare_data(df: pd.DataFrame):
    mapping = resolve_columns(df.columns)
    missing = [f for f in REQUIRED_MIN if f not in mapping]

    work = pd.DataFrame(index=df.index)
    for logical, actual in mapping.items():
        work[logical] = df[actual]

    for logical in COLUMN_ALIASES:
        if logical not in work.columns:
            work[logical] = np.nan

    work["ob_tad"] = clean_numeric(work["ob_tad"])
    work["total_visits"] = clean_numeric(work["total_visits"])
    work["ptp_amount"] = clean_numeric(work["ptp_amount"])
    work["action_date"] = clean_date(work["action_date"])
    work["endorsement_date"] = clean_date(work["endorsement_date"])
    work["ptp_date"] = clean_date(work["ptp_date"])

    for c in ["status", "substatus", "industry", "agent", "area", "area2",
              "sub_area", "active", "fv_result", "tele_field",
              "area_list", "cured_flag", "risk_level", "bal_distro", "bp_flag", "rfd"]:
        work[c] = work[c].astype(str).str.strip()
        work[c] = work[c].replace({"nan": "Unspecified", "": "Unspecified", "None": "Unspecified"})

    work["concat"] = work["concat"].astype(str).str.strip()
    work["status_norm"] = work["status"].str.title()
    work["has_payment"] = work["payment"].astype(str).str.strip().replace(
        {"nan": "", "None": ""}
    ).ne("")

    def _bom_fresh(v):
        vl = str(v).lower()
        if "bom" in vl:
            return "BOM"
        if "fresh" in vl:
            return "FRESH"
        return v

    work["bom_fresh"] = work["risk_level"].apply(_bom_fresh)

    # ---- PTP validity: a record only counts as a PTP when PTP Date is present ----
    work["is_valid_ptp"] = work["ptp_date"].notna()

    # ---- BP (Broken Promise) enhanced identification ----
    # 1) Trust an explicit Yes/True-style value already present in the "BP?" column.
    def _norm_yes(v):
        return str(v).strip().lower() in ("yes", "y", "true", "1", "bp", "broken", "broken promise")

    work["bp_raw_flag"] = work["bp_flag"].apply(_norm_yes)

    # 2) Backstop with computed logic: a valid PTP whose promised date has already
    #    lapsed (relative to the most recent Action Date in this extract, used as
    #    the "as of" reference point) with no cure recorded is also a broken promise.
    reference_date = work["action_date"].max() if work["action_date"].notna().any() else pd.Timestamp.now().normalize()
    work["ptp_lapsed"] = (
        work["is_valid_ptp"]
        & (work["ptp_date"] < reference_date)
        & (work["status_norm"] != "Cured")
    )
    work["is_bp"] = work["is_valid_ptp"] & (work["bp_raw_flag"] | work["ptp_lapsed"])
    work["bp_status"] = np.where(work["is_bp"], "Yes", "No")

    return work, mapping, missing


# ============================================================================
# AGGREGATION HELPERS
# ============================================================================

def distinct_accounts(df: pd.DataFrame) -> pd.DataFrame:
    """One row per distinct concat (dedup for portfolio-level counts)."""
    return df.drop_duplicates(subset=["concat"])


def summarize_by(df: pd.DataFrame, group_col: str, value_col: str = "ob_tad",
                  id_col: str = "concat") -> pd.DataFrame:
    """Distinct-count accounts and sum balance per group, dedup on (id, group)."""
    tmp = df.dropna(subset=[group_col])
    tmp = tmp[tmp[group_col] != "Unspecified"] if tmp[group_col].dtype == object else tmp
    dedup = tmp.drop_duplicates(subset=[id_col, group_col])
    out = dedup.groupby(group_col, dropna=False).agg(
        Accounts=(id_col, "nunique"),
        Balance=(value_col, "sum"),
    ).reset_index()
    total_accounts = df.drop_duplicates(subset=[id_col])[id_col].nunique()
    out["Pct of Accounts"] = (out["Accounts"] / total_accounts * 100) if total_accounts else 0
    return out.sort_values("Balance", ascending=False)


def status_summary(df: pd.DataFrame) -> pd.DataFrame:
    acc = distinct_accounts(df)
    rows = []
    for st_name in ["Active", "Flowed", "Cured"]:
        sub = acc[acc["status_norm"] == st_name]
        rows.append({
            "Status": st_name,
            "Accounts": sub["concat"].nunique(),
            "Balance": sub["ob_tad"].sum(),
        })
    out = pd.DataFrame(rows)
    total_acc = acc["concat"].nunique()
    total_bal = acc["ob_tad"].sum()
    out["Pct Accounts"] = (out["Accounts"] / total_acc * 100) if total_acc else 0
    out["Pct Balance"] = (out["Balance"] / total_bal * 100) if total_bal else 0
    return out


def full_status_matrix(df: pd.DataFrame, group_col: str, label_name: str) -> pd.DataFrame:
    """Generic Total/Active/Cured/Flowed count & balance matrix, with a TOTAL row.
    Column order matches: <label>, Total Count, Total Balance,
    Active Count, Active Balance, Cured Count, Cured Balance,
    Flowed Count, Flowed Balance."""
    acc = distinct_accounts(df)
    acc = acc[acc[group_col] != "Unspecified"] if acc[group_col].dtype == object else acc
    rows = []
    for val, grp in acc.groupby(group_col):
        row = {label_name: val, "Total Count": grp["concat"].nunique(),
               "Total Balance": grp["ob_tad"].sum()}
        for st_name in ["Active", "Cured", "Flowed"]:
            s = grp[grp["status_norm"] == st_name]
            row[f"{st_name} Count"] = s["concat"].nunique()
            row[f"{st_name} Balance"] = s["ob_tad"].sum()
        rows.append(row)
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out = out.sort_values("Total Balance", ascending=False)
    total = {label_name: "TOTAL"}
    for c in out.columns:
        if c != label_name:
            total[c] = out[c].sum()
    out = pd.concat([out, pd.DataFrame([total])], ignore_index=True)
    return out


def cure_summary_by(df: pd.DataFrame, group_col: str, label_name: str) -> pd.DataFrame:
    """Cured-based (Status = CURED) analysis: Total Accounts, Cured Accounts,
    Cure Rate %, Cured Balance — used for the Cured/Payment dashboard tables."""
    acc = distinct_accounts(df)
    acc = acc[acc[group_col] != "Unspecified"] if acc[group_col].dtype == object else acc
    rows = []
    for val, grp in acc.groupby(group_col):
        total_accounts = grp["concat"].nunique()
        cured = grp[grp["status_norm"] == "Cured"]
        cured_accounts = cured["concat"].nunique()
        rows.append({
            label_name: val,
            "Total Accounts": total_accounts,
            "Cured Accounts": cured_accounts,
            "Cure Rate %": (cured_accounts / total_accounts * 100) if total_accounts else 0,
            "Cured Balance": cured["ob_tad"].sum(),
        })
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    return out.sort_values("Cured Balance", ascending=False)


def agent_matrix(df: pd.DataFrame) -> pd.DataFrame:
    acc = distinct_accounts(df)
    rows = []
    for agent, grp in acc.groupby("agent"):
        row = {"Agent": agent, "Endorsed Accounts": grp["concat"].nunique(),
               "Endorsed Balance": grp["ob_tad"].sum()}
        for st_name in ["Active", "Flowed", "Cured"]:
            s = grp[grp["status_norm"] == st_name]
            row[f"{st_name} Accounts"] = s["concat"].nunique()
            row[f"{st_name} Balance"] = s["ob_tad"].sum()
        row["Cure %"] = (row.get("Cured Accounts", 0) / row["Endorsed Accounts"] * 100
                          if row["Endorsed Accounts"] else 0)
        rows.append(row)
    return pd.DataFrame(rows).sort_values("Endorsed Balance", ascending=False)


def agent_ptp_bp_matrix(df: pd.DataFrame) -> pd.DataFrame:
    """Per-agent PTP (valid PTP Date only), BP, Cured, and RFD summary, with conversion rates:
    - PTP-to-Cured Conversion Rate % = Cured Count / PTP Count * 100
    - PTP-to-BP Conversion Rate %    = BP Count / PTP Count * 100
      (what share of promises made were subsequently broken; lower is better)
    RFD = most frequent Reason for Default tagged on the agent's accounts (excluding
    Unspecified); RFD Value = total OB across that agent's RFD-tagged accounts.
    PTP RFD / BP RFD = count of that agent's RFD-tagged accounts within the PTP / BP subsets.
    PTP Accounts/Amount only count records where PTP Date is present."""
    acc = distinct_accounts(df)
    rows = []
    for agent, grp in acc.groupby("agent"):
        ptp = grp[grp["is_valid_ptp"]]
        bp = grp[grp["is_bp"]]
        cured = grp[grp["status_norm"] == "Cured"]
        rfd_tagged = grp[grp["rfd"] != "Unspecified"]
        ptp_accounts = ptp["concat"].nunique()
        bp_count = bp["concat"].nunique()
        cured_count = cured["concat"].nunique()
        top_rfd = rfd_tagged["rfd"].mode()
        rows.append({
            "Agent": agent,
            "Total Visits": grp["total_visits"].sum(),
            "PTP Accounts": ptp_accounts,
            "PTP Amount": ptp["ptp_amount"].sum(),
            "BP Count": bp_count,
            "BP OB": bp["ob_tad"].sum(),
            "Cured Count": cured_count,
            "RFD": top_rfd.iloc[0] if not top_rfd.empty else "—",
            "RFD Value": rfd_tagged["ob_tad"].sum(),
            "PTP RFD": ptp[ptp["rfd"] != "Unspecified"]["concat"].nunique(),
            "BP RFD": bp[bp["rfd"] != "Unspecified"]["concat"].nunique(),
            "PTP-to-Cured Conversion Rate %": (cured_count / ptp_accounts * 100) if ptp_accounts else 0,
            "PTP-to-BP Conversion Rate %": (bp_count / ptp_accounts * 100) if ptp_accounts else np.nan,
        })
    return pd.DataFrame(rows).sort_values("PTP Amount", ascending=False)


def rfd_summary_by(df: pd.DataFrame, subset: str = "all") -> pd.DataFrame:
    """RFD (Reason for Default) breakdown. subset: 'all', 'ptp' (valid PTP only),
    or 'bp' (Broken Promise only)."""
    acc = distinct_accounts(df)
    if subset == "ptp":
        acc = acc[acc["is_valid_ptp"]]
    elif subset == "bp":
        acc = acc[acc["is_bp"]]
    acc = acc[acc["rfd"] != "Unspecified"]
    if acc.empty:
        return pd.DataFrame(columns=["RFD", "Count", "Value"])
    out = acc.groupby("rfd").agg(
        Count=("concat", "nunique"), Value=("ob_tad", "sum")
    ).reset_index().rename(columns={"rfd": "RFD"})
    return out.sort_values("Value", ascending=False)


def bp_summary_by(df: pd.DataFrame, group_col: str, label_name: str) -> pd.DataFrame:
    """BP Count and BP OB grouped by an arbitrary column, restricted to BP=Yes records."""
    acc = distinct_accounts(df)
    acc = acc[acc[group_col] != "Unspecified"] if acc[group_col].dtype == object else acc
    bp = acc[acc["is_bp"]]
    if bp.empty:
        return pd.DataFrame(columns=[label_name, "BP Count", "BP OB"])
    out = bp.groupby(group_col).agg(
        **{"BP Count": ("concat", "nunique"), "BP OB": ("ob_tad", "sum")}
    ).reset_index().rename(columns={group_col: label_name})
    return out.sort_values("BP OB", ascending=False)


# ============================================================================
# GEO / MAP HELPERS (Philippine provinces, for the Area List map dashboard)
# ============================================================================

PH_PROVINCE_COORDS = {
    "NCR": (14.5995, 120.9842), "METRO MANILA": (14.5995, 120.9842),
    "NATIONAL CAPITAL REGION": (14.5995, 120.9842),
    "ABRA": (17.5951, 120.7983), "AGUSAN DEL NORTE": (8.9450, 125.5319),
    "AGUSAN DEL SUR": (8.1661, 125.9528), "AKLAN": (11.8166, 122.0942),
    "ALBAY": (13.1775, 123.5280), "ANTIQUE": (11.3357, 122.0602),
    "APAYAO": (18.0085, 121.1651), "AURORA": (15.7594, 121.5591),
    "BASILAN": (6.4221, 121.9690), "BATAAN": (14.6417, 120.4818),
    "BATANES": (20.4487, 121.9702), "BATANGAS": (13.9294, 121.1637),
    "BENGUET": (16.4023, 120.5960), "BILIRAN": (11.5836, 124.4645),
    "BOHOL": (9.8500, 124.1435), "BUKIDNON": (8.0515, 125.0985),
    "BULACAN": (14.7943, 120.8792), "CAGAYAN": (17.9989, 121.7534),
    "CAMARINES NORTE": (14.1391, 122.7573), "CAMARINES SUR": (13.6252, 123.1829),
    "CAMIGUIN": (9.1736, 124.7300), "CAPIZ": (11.3889, 122.6277),
    "CATANDUANES": (13.7089, 124.2422), "CAVITE": (14.2456, 120.8786),
    "CEBU": (10.3157, 123.8854), "COTABATO": (7.2072, 124.2422),
    "DAVAO DE ORO": (7.6167, 126.1667), "COMPOSTELA VALLEY": (7.6167, 126.1667),
    "DAVAO DEL NORTE": (7.5619, 125.6549), "DAVAO DEL SUR": (6.7656, 125.3284),
    "DAVAO OCCIDENTAL": (6.1000, 125.6000), "DAVAO ORIENTAL": (7.3172, 126.5420),
    "DINAGAT ISLANDS": (10.1281, 125.6094), "EASTERN SAMAR": (11.6067, 125.5000),
    "GUIMARAS": (10.5928, 122.6325), "IFUGAO": (16.8300, 121.1710),
    "ILOCOS NORTE": (18.1647, 120.7116), "ILOCOS SUR": (17.5755, 120.3869),
    "ILOILO": (10.7202, 122.5621), "ISABELA": (17.0000, 121.8333),
    "KALINGA": (17.4766, 121.3521), "LA UNION": (16.6159, 120.3209),
    "LAGUNA": (14.2691, 121.4113), "LANAO DEL NORTE": (8.1156, 123.9210),
    "LANAO DEL SUR": (7.8232, 124.4357), "LEYTE": (10.8731, 124.8811),
    "MAGUINDANAO": (6.9423, 124.4198), "MAGUINDANAO DEL NORTE": (7.1800, 124.4500),
    "MAGUINDANAO DEL SUR": (6.9000, 124.4000),
    "MARINDUQUE": (13.4771, 121.9032), "MASBATE": (12.3686, 123.6417),
    "MISAMIS OCCIDENTAL": (8.3375, 123.7071), "MISAMIS ORIENTAL": (8.6109, 124.7739),
    "MOUNTAIN PROVINCE": (17.0417, 121.1087), "NEGROS OCCIDENTAL": (10.4275, 122.9847),
    "NEGROS ORIENTAL": (9.6168, 123.0113), "NORTHERN SAMAR": (12.4700, 124.6400),
    "NUEVA ECIJA": (15.5784, 121.0687), "NUEVA VIZCAYA": (16.3301, 121.1710),
    "OCCIDENTAL MINDORO": (13.1024, 120.7651), "ORIENTAL MINDORO": (13.0565, 121.4069),
    "PALAWAN": (9.8349, 118.7384), "PAMPANGA": (15.0794, 120.6200),
    "PANGASINAN": (15.8949, 120.2863),
    "QUEZON": (14.0313, 122.1106),           # Quezon Province
    "QUEZON PROVINCE": (14.0313, 122.1106),
    "QUEZON CITY": (14.6760, 121.0437),      # Quezon City (Metro Manila) — kept distinct from Quezon Province
    "QUIRINO": (16.3676, 121.5479), "RIZAL": (14.6037, 121.3084),
    "ROMBLON": (12.5778, 122.2695), "SAMAR": (11.9804, 124.9944),
    "SARANGANI": (5.9591, 125.2228), "SIQUIJOR": (9.1911, 123.5952),
    "SORSOGON": (12.9743, 124.0150), "SOUTH COTABATO": (6.2969, 124.8511),
    "SOUTHERN LEYTE": (10.3365, 125.1717), "SULTAN KUDARAT": (6.5069, 124.4169),
    "SULU": (6.0474, 121.0024), "SURIGAO DEL NORTE": (9.7899, 125.4947),
    "SURIGAO DEL SUR": (8.7512, 126.1378), "TARLAC": (15.4802, 120.5979),
    "TAWI-TAWI": (5.1339, 119.9552), "ZAMBALES": (15.5082, 120.0691),
    "ZAMBOANGA DEL NORTE": (8.1527, 123.2577), "ZAMBOANGA DEL SUR": (7.8383, 123.2984),
    "ZAMBOANGA SIBUGAY": (7.5222, 122.8198),
    # ---- Additional Metro Manila cities / districts and CAR region ----
    "PASIG": (14.5764, 121.0851),
    "CORDILLERA": (16.4023, 120.5960),               # CAR region (Baguio-area centroid)
    "CORDILLERA ADMINISTRATIVE REGION": (16.4023, 120.5960),
    "PARANAQUE": (14.4793, 121.0198),
    "MAKATI": (14.5547, 121.0244),
    "MARIKINA": (14.6507, 121.1029),
    "BICUTAN": (14.5027, 121.0509),
    "TAGUIG": (14.5176, 121.0509),
    "VALENZUELA CITY": (14.7000, 120.9830),
    "VALENZUELA": (14.7000, 120.9830),
    "MALABON": (14.6650, 120.9567),
}


def _strip_accents(text: str) -> str:
    """Remove accents/diacritics (e.g. ñ -> n) for robust text matching."""
    return "".join(c for c in unicodedata.normalize("NFKD", text) if not unicodedata.combining(c))


def geocode_area(name: str):
    """Best-effort match of a free-text area/province/city label to PH lat/lon.
    Uses exact match first, then longest-name-first substring matching so more
    specific labels (e.g. 'QUEZON CITY') are preferred over shorter, broader
    ones (e.g. 'QUEZON') when both could match the input text."""
    if not name or str(name).strip() in ("", "nan", "None", "Unspecified"):
        return None
    key = re.sub(r"\s+", " ", _strip_accents(str(name).strip()).upper())
    if key in PH_PROVINCE_COORDS:
        return PH_PROVINCE_COORDS[key]
    for province in sorted(PH_PROVINCE_COORDS.keys(), key=len, reverse=True):
        if province in key or key in province:
            return PH_PROVINCE_COORDS[province]
    return None


def style_and_write_sheet(ws, header_fill_hex="0057B8"):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    header_fill = PatternFill("solid", fgColor=header_fill_hex)
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    for i, col_cells in enumerate(ws.columns, start=1):
        header_val = str(col_cells[0].value) if col_cells[0].value is not None else ""
        max_len = max([len(str(c.value)) if c.value is not None else 0 for c in col_cells] + [len(header_val)])
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 42)
        if "balance" in header_val.lower() or header_val.lower() in ("ob", "ob/tad"):
            for c in col_cells[1:]:
                if isinstance(c.value, (int, float)):
                    c.number_format = '"₱"#,##0.00'
        if i % 2 == 0:
            for r_idx, c in enumerate(col_cells[1:], start=2):
                if r_idx % 2 == 0:
                    from openpyxl.styles import PatternFill as PF
                    c.fill = PF("solid", fgColor="EAF3FF")


def to_excel_bytes(sheets: dict, title="SBC B2 BEL Collection Dashboard") -> bytes:
    from openpyxl.styles import Font, PatternFill, Alignment

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        wb = writer.book
        info_ws = wb.create_sheet("Report Info", 0)
        info_ws["B2"] = title
        info_ws["B2"].font = Font(size=20, bold=True, color="FFFFFF")
        info_ws["B2"].fill = PatternFill("solid", fgColor="0057B8")
        info_ws.merge_cells("B2:G4")
        for row in info_ws["B2:G4"]:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="0057B8")
        info_ws["B6"] = "Security Bank Collections — Executive Report"
        info_ws["B6"].font = Font(size=12, italic=True, color="003B7A")
        info_ws["B7"] = f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}"
        info_ws["B7"].font = Font(size=11, color="333333")
        info_ws.column_dimensions["A"].width = 3

        for name, d in sheets.items():
            d.to_excel(writer, sheet_name=name[:31], index=False)

        for name in sheets:
            ws = wb[name[:31]]
            style_and_write_sheet(ws)

    return buf.getvalue()


# ============================================================================
# SIDEBAR — FILE UPLOAD
# ============================================================================

st.markdown(
    f"""
    <div class="sbc-header">
        <div class="sbc-logo">SB</div>
        <div class="sbc-header-text">
            <h1>SBC B2 BEL Collection Dashboard</h1>
            <p>Executive Collections &amp; Curing Performance Monitoring · Security Bank</p>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

st.sidebar.title("📁 Data Source")
uploaded = st.sidebar.file_uploader(
    "Upload collections file",
    type=["xlsx", "xls", "csv"],
    help="Upload the Field Collections monitoring extract (XLSX, XLS, or CSV).",
)

if not uploaded:
    st.info(
        "👈 Upload an XLSX, XLS, or CSV file from the sidebar to begin. "
        "The dashboard supports large datasets (100,000+ rows) and will "
        "auto-detect columns such as **concat**, **OB/TAD**, **Status**, "
        "**Industry**, **Agent**, **Area**, and **FV Result**."
    )
    st.stop()

raw_bytes = uploaded.getvalue()
with st.spinner("Loading file..."):
    raw_df = load_file(raw_bytes, uploaded.name)

with st.spinner("Processing and validating data..."):
    data, colmap, missing_fields = prepare_data(raw_df)

if missing_fields:
    st.sidebar.warning(
        "Some key columns were not auto-detected: "
        + ", ".join(missing_fields)
        + ". Related metrics may show as zero/Unspecified."
    )

st.sidebar.success(f"Loaded {len(raw_df):,} rows · {distinct_accounts(data)['concat'].nunique():,} distinct accounts")

with st.sidebar.expander("🔎 Detected column mapping"):
    st.dataframe(pd.DataFrame(
        [{"Field": k, "Source Column": v} for k, v in colmap.items()]
    ), hide_index=True, use_container_width=True)

# ============================================================================
# SIDEBAR — FILTERS
# ============================================================================

st.sidebar.markdown("---")
st.sidebar.title("🎛️ Filters")

if st.sidebar.button("🔄 Clear all filters"):
    for k in list(st.session_state.keys()):
        if k.startswith("flt_"):
            del st.session_state[k]
    st.rerun()


def multiselect_filter(label, col, key):
    opts = sorted([o for o in data[col].dropna().unique() if o != ""])
    return st.sidebar.multiselect(label, opts, key=f"flt_{key}")


f_status = multiselect_filter("Status", "status", "status")
f_substatus = multiselect_filter("Tele Result (SubStatus)", "substatus", "substatus")
f_industry = multiselect_filter("Industry", "industry", "industry")
f_area = multiselect_filter("Area", "area", "area")
f_area2 = multiselect_filter("Area 2 / Area Break", "area2", "area2")
f_agent = multiselect_filter("Agent", "agent", "agent")
f_active = multiselect_filter("Active Status", "active", "active")
f_fv = multiselect_filter("Field Result (Fv_Result)", "fv_result", "fv")
f_product = multiselect_filter("Product Type (Tele/Field)", "tele_field", "product")
f_riskclass = multiselect_filter("BOM / FRESH", "bom_fresh", "bomfresh")
f_bp = multiselect_filter("BP (Broken Promise) Status", "bp_status", "bpstatus")

min_end = data["endorsement_date"].min()
max_end = data["endorsement_date"].max()
f_end_date = None
if pd.notna(min_end) and pd.notna(max_end):
    f_end_date = st.sidebar.date_input(
        "Date of Endorsement range", value=(min_end.date(), max_end.date()),
        min_value=min_end.date(), max_value=max_end.date(), key="flt_enddate",
    )

min_act = data["action_date"].min()
max_act = data["action_date"].max()
f_act_date = None
if pd.notna(min_act) and pd.notna(max_act):
    f_act_date = st.sidebar.date_input(
        "Action Date range", value=(min_act.date(), max_act.date()),
        min_value=min_act.date(), max_value=max_act.date(), key="flt_actdate",
    )

min_ptp = data["ptp_date"].min()
max_ptp = data["ptp_date"].max()
f_ptp_date = None
if pd.notna(min_ptp) and pd.notna(max_ptp):
    f_ptp_date = st.sidebar.date_input(
        "PTP Date range", value=(min_ptp.date(), max_ptp.date()),
        min_value=min_ptp.date(), max_value=max_ptp.date(), key="flt_ptpdate",
    )

filtered = data.copy()
if f_status:
    filtered = filtered[filtered["status"].isin(f_status)]
if f_substatus:
    filtered = filtered[filtered["substatus"].isin(f_substatus)]
if f_industry:
    filtered = filtered[filtered["industry"].isin(f_industry)]
if f_area:
    filtered = filtered[filtered["area"].isin(f_area)]
if f_area2:
    filtered = filtered[filtered["area2"].isin(f_area2)]
if f_agent:
    filtered = filtered[filtered["agent"].isin(f_agent)]
if f_active:
    filtered = filtered[filtered["active"].isin(f_active)]
if f_fv:
    filtered = filtered[filtered["fv_result"].isin(f_fv)]
if f_product:
    filtered = filtered[filtered["tele_field"].isin(f_product)]
if f_riskclass:
    filtered = filtered[filtered["bom_fresh"].isin(f_riskclass)]
if f_bp:
    filtered = filtered[filtered["bp_status"].isin(f_bp)]
if f_end_date and len(f_end_date) == 2:
    start, end = pd.Timestamp(f_end_date[0]), pd.Timestamp(f_end_date[1])
    filtered = filtered[
        filtered["endorsement_date"].isna()
        | filtered["endorsement_date"].between(start, end)
    ]
if f_act_date and len(f_act_date) == 2:
    start, end = pd.Timestamp(f_act_date[0]), pd.Timestamp(f_act_date[1])
    filtered = filtered[
        filtered["action_date"].isna()
        | filtered["action_date"].between(start, end)
    ]
if f_ptp_date and len(f_ptp_date) == 2:
    start, end = pd.Timestamp(f_ptp_date[0]), pd.Timestamp(f_ptp_date[1])
    filtered = filtered[
        filtered["ptp_date"].isna()
        | filtered["ptp_date"].between(start, end)
    ]

if filtered.empty:
    st.warning("No records match the current filter selection.")
    st.stop()

st.caption(
    f"Data as of {datetime.now().strftime('%B %d, %Y %I:%M %p')} · "
    f"Showing {distinct_accounts(filtered)['concat'].nunique():,} of "
    f"{distinct_accounts(data)['concat'].nunique():,} distinct accounts"
)

tabs = st.tabs([
    "🏦 Executive Summary", "🏭 Industry", "🧑‍💼 Agent", "📞 Tele & Field Results",
    "🗺️ Area", "🌍 Area Map", "📶 Active Status", "💚 Cured / Payment",
    "📋 Account Details", "⬇️ Downloads",
])

# ============================================================================
# TAB 1 — EXECUTIVE SUMMARY
# ============================================================================

with tabs[0]:
    acc = distinct_accounts(filtered)
    total_accounts = acc["concat"].nunique()
    total_balance = acc["ob_tad"].sum()

    def bucket_status(status_name):
        s = acc[acc["status_norm"] == status_name]
        return s["concat"].nunique(), s["ob_tad"].sum()

    active_acc, active_bal = bucket_status("Active")
    flowed_acc, flowed_bal = bucket_status("Flowed")
    cured_acc, cured_bal = bucket_status("Cured")
    cure_rate = (cured_acc / total_accounts * 100) if total_accounts else 0
    cure_bal_rate = (cured_bal / total_balance * 100) if total_balance else 0

    st.subheader("Executive KPIs")
    r1 = st.columns(5)
    r1[0].metric("Total Endorsed Accounts", fnum(total_accounts))
    r1[1].metric("Total Outstanding Balance", PHP(total_balance))
    r1[2].metric("Active Accounts", fnum(active_acc))
    r1[3].metric("Active Balance", PHP(active_bal))
    r1[4].metric("Flowed Accounts", fnum(flowed_acc))

    r2 = st.columns(5)
    r2[0].metric("Flowed Balance", PHP(flowed_bal))
    r2[1].metric("Cured Accounts", fnum(cured_acc))
    r2[2].metric("Cured Balance", PHP(cured_bal))
    r2[3].metric("Cure Rate", pct(cure_rate))
    r2[4].metric("Cure Balance Rate", pct(cure_bal_rate))

    st.markdown("### Status Distribution")
    ss = status_summary(filtered)
    c1, c2, c3 = st.columns(3)
    with c1:
        fig = px.pie(ss, names="Status", values="Accounts", title="Accounts by Status",
                     color="Status", color_discrete_map=STATUS_COLORS, hole=0)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_1")
    with c2:
        fig = px.pie(ss, names="Status", values="Balance", title="Balance by Status (Donut)",
                     color="Status", color_discrete_map=STATUS_COLORS, hole=0.5)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_2")
    with c3:
        m = ss.melt(id_vars="Status", value_vars=["Accounts", "Balance"],
                     var_name="Metric", value_name="Value")
        fig = px.bar(m, x="Status", y="Value", color="Status", facet_col="Metric",
                     color_discrete_map=STATUS_COLORS, title="Accounts & Balance by Status")
        fig.update_yaxes(matches=None)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_3")

    st.dataframe(
        ss.style.format({"Balance": PHP, "Pct Accounts": "{:.1f}%", "Pct Balance": "{:.1f}%"}),
        use_container_width=True, hide_index=True,
    )

    # ---- BOM / FRESH (renamed from "Risk Level") ----
    st.markdown("---")
    st.markdown("### 🏷️ BOM / FRESH")
    bomfresh = summarize_by(filtered, "bom_fresh").rename(columns={"bom_fresh": "BOM/FRESH"})
    if not bomfresh.empty:
        bom_row = bomfresh[bomfresh["BOM/FRESH"] == "BOM"]
        fresh_row = bomfresh[bomfresh["BOM/FRESH"] == "FRESH"]
        total_bom_count = int(bom_row["Accounts"].sum())
        total_bom_ob = bom_row["Balance"].sum()
        total_fresh_count = int(fresh_row["Accounts"].sum())
        total_fresh_ob = fresh_row["Balance"].sum()

        r3 = st.columns(4)
        r3[0].metric("Total BOM Count", fnum(total_bom_count))
        r3[1].metric("Total BOM OB", PHP(total_bom_ob))
        r3[2].metric("Total FRESH Count", fnum(total_fresh_count))
        r3[3].metric("Total FRESH OB", PHP(total_fresh_ob))

        c1, c2, c3 = st.columns(3)
        with c1:
            fig = px.bar(bomfresh, x="BOM/FRESH", y="Accounts", color="BOM/FRESH",
                         title="BOM vs FRESH — Account Count",
                         color_discrete_sequence=[SB_ORANGE, SB_BLUE])
            st.plotly_chart(fig, use_container_width=True, key=f"chart_4")
        with c2:
            fig = px.pie(bomfresh, names="BOM/FRESH", values="Accounts", hole=0.5,
                         title="BOM vs FRESH — Donut",
                         color_discrete_sequence=[SB_ORANGE, SB_BLUE])
            st.plotly_chart(fig, use_container_width=True, key=f"chart_5")
        with c3:
            fig = px.bar(bomfresh, x="BOM/FRESH", y="Balance", color="BOM/FRESH",
                         title="BOM vs FRESH — OB Distribution",
                         color_discrete_sequence=[SB_ORANGE, SB_BLUE])
            st.plotly_chart(fig, use_container_width=True, key=f"chart_6")
        st.dataframe(bomfresh.style.format({"Balance": PHP, "Pct of Accounts": "{:.1f}%"}),
                     use_container_width=True, hide_index=True)
    else:
        st.info("No BOM/FRESH values detected in this dataset.")

    # ---- Balance Distribution (BAL Distro) ----
    st.markdown("---")
    st.markdown("### 💵 Balance Distribution (BAL Distro)")
    bal_mat = full_status_matrix(filtered, "bal_distro", "BAL Distro")
    if not bal_mat.empty:
        money_cols = [c for c in bal_mat.columns if "Balance" in c]
        st.dataframe(bal_mat.style.format({c: PHP for c in money_cols}),
                     use_container_width=True, hide_index=True)
        bal_body = bal_mat[bal_mat["BAL Distro"] != "TOTAL"]
        c1, c2 = st.columns(2)
        with c1:
            fig = px.bar(bal_body.sort_values("Total Count"),
                         x="Total Count", y="BAL Distro", orientation="h",
                         title="Balance Distribution — Accounts by Bucket",
                         color="Total Count", color_continuous_scale=BLUES_SCALE)
            st.plotly_chart(fig, use_container_width=True, key=f"chart_7")
        with c2:
            long = bal_body[["BAL Distro", "Active Count", "Cured Count", "Flowed Count"]].melt(
                id_vars="BAL Distro", value_vars=["Active Count", "Cured Count", "Flowed Count"],
                var_name="Status", value_name="Accounts",
            )
            long["Status"] = long["Status"].str.replace(" Count", "")
            fig = px.bar(long, x="BAL Distro", y="Accounts", color="Status", barmode="stack",
                         color_discrete_map=STATUS_COLORS, title="Balance Distribution by Status")
            st.plotly_chart(fig, use_container_width=True, key=f"chart_8")
        fig = px.treemap(bal_body, path=["BAL Distro"], values="Total Balance",
                          color="Total Balance", color_continuous_scale=BLUES_SCALE,
                          title="Treemap by BAL Distro")
        st.plotly_chart(fig, use_container_width=True, key=f"chart_9")
    else:
        st.info("No BAL Distro values detected in this dataset.")

# ============================================================================
# TAB 2 — INDUSTRY DASHBOARD
# ============================================================================

with tabs[1]:
    st.subheader("Industry Performance Matrix")
    imat = full_status_matrix(filtered, "industry", "INDUSTRY")
    body = imat[imat["INDUSTRY"] != "TOTAL"]
    cure_ind = cure_summary_by(filtered, "industry", "Industry")

    st.markdown("#### 📌 Industry KPIs")
    if not body.empty and not cure_ind.empty:
        top_bal_row = body.loc[body["Total Balance"].idxmax()]
        top_cure_row = cure_ind.loc[cure_ind["Cure Rate %"].idxmax()]
        low_cure_row = cure_ind.loc[cure_ind["Cure Rate %"].idxmin()]
        avg_cure_rate = cure_ind["Cure Rate %"].mean()
        below_avg_count = int((cure_ind["Cure Rate %"] < avg_cure_rate).sum())

        k1, k2, k3, k4, k5 = st.columns(5)
        k1.metric("Total Industries", fnum(body["INDUSTRY"].nunique()))
        k2.metric("Top Industry by Balance", top_bal_row["INDUSTRY"], PHP(top_bal_row["Total Balance"]))
        k3.metric("Best Cure Rate", top_cure_row["Industry"], pct(top_cure_row["Cure Rate %"]))
        k4.metric("⚠️ Needs Attention (Lowest Cure Rate)", low_cure_row["Industry"],
                   pct(low_cure_row["Cure Rate %"]), delta_color="inverse")
        k5.metric("Industries Below Avg Cure Rate", fnum(below_avg_count), f"Avg: {pct(avg_cure_rate)}",
                   delta_color="off")
    else:
        st.info("Not enough Industry data in the current filter selection to compute KPIs.")

    money_cols = [c for c in imat.columns if "Balance" in c]
    st.dataframe(
        imat.style.format({c: PHP for c in money_cols}),
        use_container_width=True, hide_index=True, height=420,
    )

    c1, c2 = st.columns(2)
    with c1:
        fig = px.bar(body.sort_values("Total Balance", ascending=True),
                     x="Total Balance", y="INDUSTRY", orientation="h",
                     title="Industry Balance Ranking", color="Total Balance",
                     color_continuous_scale=BLUES_SCALE)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_10")
    with c2:
        long = body[["INDUSTRY", "Active Count", "Cured Count", "Flowed Count"]].melt(
            id_vars="INDUSTRY", value_vars=["Active Count", "Cured Count", "Flowed Count"],
            var_name="Status", value_name="Accounts",
        )
        long["Status"] = long["Status"].str.replace(" Count", "")
        fig = px.bar(long, x="INDUSTRY", y="Accounts", color="Status",
                     color_discrete_map=STATUS_COLORS, title="Industry Status Distribution")
        st.plotly_chart(fig, use_container_width=True, key=f"chart_11")

    fig = px.treemap(body, path=["INDUSTRY"], values="Total Balance",
                      color="Total Balance", color_continuous_scale=BLUES_SCALE,
                      title="Industry Treemap (by Balance)")
    st.plotly_chart(fig, use_container_width=True, key=f"chart_12")

    st.markdown("### 💚 Cured Accounts & Balance by Industry")
    c1, c2 = st.columns(2)
    with c1:
        fig = px.bar(cure_ind.sort_values("Cured Accounts"), x="Cured Accounts", y="Industry",
                     orientation="h", title="Cured Accounts by Industry", color="Cured Accounts",
                     color_continuous_scale=GREENS_SCALE)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_13")
    with c2:
        fig = px.bar(cure_ind.sort_values("Cured Balance"), x="Cured Balance", y="Industry",
                     orientation="h", title="Cured Balance by Industry", color="Cured Balance",
                     color_continuous_scale=GREENS_SCALE)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_14")

    st.markdown("### 🏆 Industry Cure Rate Ranking")
    cure_ind_ranked = cure_ind.sort_values("Cure Rate %", ascending=True)
    fig = px.bar(cure_ind_ranked, x="Cure Rate %", y="Industry", orientation="h",
                 title="Industry Cure Rate Ranking", color="Cure Rate %",
                 color_continuous_scale=GREENS_SCALE)
    st.plotly_chart(fig, use_container_width=True, key=f"chart_15")

    st.markdown("### 🔎 Drill-Down: Industry → Tele Result / Field Result")
    industry_options = sorted([i for i in body["INDUSTRY"].unique()])
    if industry_options:
        sel_industry = st.selectbox("Select an Industry to drill into", industry_options, key="industry_drilldown")
        ind_slice = filtered[filtered["industry"] == sel_industry]
        c1, c2 = st.columns(2)
        with c1:
            tele_dist = summarize_by(ind_slice, "substatus").rename(columns={"substatus": "Tele Result"})
            fig = px.bar(tele_dist.sort_values("Accounts"), x="Accounts", y="Tele Result",
                         orientation="h", title=f"{sel_industry} → Tele Result Distribution",
                         color="Accounts", color_continuous_scale=BLUES_SCALE)
            st.plotly_chart(fig, use_container_width=True, key=f"chart_16")
        with c2:
            field_dist = summarize_by(ind_slice, "fv_result").rename(columns={"fv_result": "Field Result"})
            fig = px.bar(field_dist.sort_values("Accounts"), x="Accounts", y="Field Result",
                         orientation="h", title=f"{sel_industry} → Field Result Distribution",
                         color="Accounts", color_continuous_scale=GREENS_SCALE)
            st.plotly_chart(fig, use_container_width=True, key=f"chart_17")

    st.markdown("### Industry vs Tele Result / Field Result (Overall)")
    c1, c2 = st.columns(2)
    with c1:
        tele_by_ind = filtered.dropna(subset=["industry", "substatus"])
        tele_by_ind = tele_by_ind[(tele_by_ind["industry"] != "Unspecified") & (tele_by_ind["substatus"] != "Unspecified")]
        pivot_tele = distinct_accounts(tele_by_ind).pivot_table(
            index="industry", columns="substatus", values="concat", aggfunc="nunique", fill_value=0)
        top_tele_cols = pivot_tele.sum(axis=0).sort_values(ascending=False).head(8).index
        m = pivot_tele[top_tele_cols].reset_index().melt(id_vars="industry", var_name="Tele Result", value_name="Accounts")
        fig = px.bar(m, x="industry", y="Accounts", color="Tele Result",
                     title="Industry vs Tele Result (Top Categories)")
        st.plotly_chart(fig, use_container_width=True, key=f"chart_18")
    with c2:
        field_by_ind = filtered.dropna(subset=["industry", "fv_result"])
        field_by_ind = field_by_ind[(field_by_ind["industry"] != "Unspecified") & (field_by_ind["fv_result"] != "Unspecified")]
        pivot_field = distinct_accounts(field_by_ind).pivot_table(
            index="industry", columns="fv_result", values="concat", aggfunc="nunique", fill_value=0)
        top_field_cols = pivot_field.sum(axis=0).sort_values(ascending=False).head(8).index
        m = pivot_field[top_field_cols].reset_index().melt(id_vars="industry", var_name="Field Result", value_name="Accounts")
        fig = px.bar(m, x="industry", y="Accounts", color="Field Result",
                     title="Industry vs Field Result (Top Categories)")
        st.plotly_chart(fig, use_container_width=True, key=f"chart_19")

# ============================================================================
# TAB 3 — AGENT DASHBOARD
# ============================================================================

with tabs[2]:
    st.subheader("Agent Performance")
    amat = agent_matrix(filtered)
    ptp_bp_agent = agent_ptp_bp_matrix(filtered)

    st.markdown("#### 📌 Agent KPIs")
    if not amat.empty:
        top_bal_row = amat.loc[amat["Endorsed Balance"].idxmax()]
        top_cure_row = amat.loc[amat["Cure %"].idxmax()]
        k1, k2, k3, k4, k5 = st.columns(5)
        k1.metric("Total Agents", fnum(amat["Agent"].nunique()))
        k2.metric("Top Agent by Balance", top_bal_row["Agent"], PHP(top_bal_row["Endorsed Balance"]))
        k3.metric("Top Agent by Cure Rate", top_cure_row["Agent"], pct(top_cure_row["Cure %"]))
        if not ptp_bp_agent.empty and ptp_bp_agent["PTP Amount"].sum() > 0:
            top_ptp_row = ptp_bp_agent.loc[ptp_bp_agent["PTP Amount"].idxmax()]
            k4.metric("Top Agent by PTP Amount", top_ptp_row["Agent"], PHP(top_ptp_row["PTP Amount"]))
        else:
            k4.metric("Top Agent by PTP Amount", "—", "No valid PTPs")
        if not ptp_bp_agent.empty and ptp_bp_agent["BP Count"].sum() > 0:
            top_bp_row = ptp_bp_agent.loc[ptp_bp_agent["BP Count"].idxmax()]
            k5.metric("⚠️ Highest BP Count (Needs Coaching)", top_bp_row["Agent"],
                       f"{fnum(top_bp_row['BP Count'])} BP", delta_color="inverse")
        else:
            k5.metric("⚠️ Highest BP Count (Needs Coaching)", "—", "No BPs", delta_color="off")
    else:
        st.info("Not enough Agent data in the current filter selection to compute KPIs.")

    st.markdown("#### 📊 Productivity & Conversion KPIs")
    agent_acc = distinct_accounts(filtered)
    total_visits_agent = agent_acc["total_visits"].sum()
    total_bp = ptp_bp_agent["BP Count"].sum() if not ptp_bp_agent.empty else 0
    total_ptp = ptp_bp_agent["PTP Accounts"].sum() if not ptp_bp_agent.empty else 0
    total_cured_agent = ptp_bp_agent["Cured Count"].sum() if not ptp_bp_agent.empty else 0
    ptp_to_bp_rate = (total_bp / total_ptp * 100) if total_ptp else 0
    ptp_to_cured_rate = (total_cured_agent / total_ptp * 100) if total_ptp else 0

    j1, j2, j3, j4 = st.columns(4)
    j1.metric("Total Visits", fnum(total_visits_agent))
    j2.metric("BP Count", fnum(total_bp))
    j3.metric("PTP Count", fnum(total_ptp))
    j4.metric("Cured Count", fnum(total_cured_agent))
    j5, j6 = st.columns(2)
    render_rate_metric(j5, "PTP-to-BP Conversion Rate", pct(ptp_to_bp_rate), ptp_to_bp_rate,
                        good=15, warn=30, higher_is_better=False,
                        help_text="BP Count ÷ PTP Count × 100 — share of promises that were "
                                   "later broken. Lower is better.")
    render_rate_metric(j6, "PTP-to-Cured Conversion Rate", pct(ptp_to_cured_rate), ptp_to_cured_rate,
                        good=70, warn=40)

    money_cols = [c for c in amat.columns if "Balance" in c]
    st.dataframe(
        amat.style.format({**{c: PHP for c in money_cols}, "Cure %": "{:.1f}%"}),
        use_container_width=True, hide_index=True, height=420,
    )

    top20 = amat.head(20)
    c1, c2, c3 = st.columns(3)
    with c1:
        fig = px.bar(top20.sort_values("Endorsed Balance"), x="Endorsed Balance", y="Agent",
                     orientation="h", title="Top 20 Agents by Balance",
                     color="Endorsed Balance", color_continuous_scale=BLUES_SCALE)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_20")
    with c2:
        t = amat.sort_values("Cure %", ascending=False).head(20)
        fig = px.bar(t.sort_values("Cure %"), x="Cure %", y="Agent", orientation="h",
                     title="Top 20 Agents by Cure Rate", color="Cure %",
                     color_continuous_scale=GREENS_SCALE)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_21")
    with c3:
        t = amat.sort_values("Endorsed Accounts", ascending=False).head(20)
        fig = px.bar(t.sort_values("Endorsed Accounts"), x="Endorsed Accounts", y="Agent",
                     orientation="h", title="Top 20 Agents by # of Accounts",
                     color="Endorsed Accounts", color_continuous_scale=BLUES_SCALE)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_22")

    st.markdown("### 🏆 Agent Leaderboards")
    lb_tabs = st.tabs(["Highest Accounts", "Highest OB", "Highest Cured Count",
                       "Highest Cured Balance", "Highest Cure Rate %"])
    with lb_tabs[0]:
        lb = amat.sort_values("Endorsed Accounts", ascending=False).reset_index(drop=True)
        lb.insert(0, "Rank", lb.index + 1)
        st.dataframe(
            lb[["Rank", "Agent", "Endorsed Accounts", "Endorsed Balance"]].head(15)
            .style.format({"Endorsed Balance": PHP}),
            use_container_width=True, hide_index=True,
        )
    with lb_tabs[1]:
        lb = amat.sort_values("Endorsed Balance", ascending=False).reset_index(drop=True)
        lb.insert(0, "Rank", lb.index + 1)
        st.dataframe(
            lb[["Rank", "Agent", "Endorsed Balance", "Endorsed Accounts"]].head(15)
            .style.format({"Endorsed Balance": PHP}),
            use_container_width=True, hide_index=True,
        )
    with lb_tabs[2]:
        lb = amat.sort_values("Cured Accounts", ascending=False).reset_index(drop=True)
        lb.insert(0, "Rank", lb.index + 1)
        st.dataframe(
            lb[["Rank", "Agent", "Cured Accounts", "Cured Balance", "Cure %"]].head(15)
            .style.format({"Cured Balance": PHP, "Cure %": "{:.1f}%"}),
            use_container_width=True, hide_index=True,
        )
    with lb_tabs[3]:
        lb = amat.sort_values("Cured Balance", ascending=False).reset_index(drop=True)
        lb.insert(0, "Rank", lb.index + 1)
        st.dataframe(
            lb[["Rank", "Agent", "Cured Balance", "Cured Accounts", "Cure %"]].head(15)
            .style.format({"Cured Balance": PHP, "Cure %": "{:.1f}%"}),
            use_container_width=True, hide_index=True,
        )
    with lb_tabs[4]:
        lb = amat.sort_values("Cure %", ascending=False).reset_index(drop=True)
        lb.insert(0, "Rank", lb.index + 1)
        st.dataframe(
            lb[["Rank", "Agent", "Cure %", "Cured Accounts", "Endorsed Accounts"]].head(15)
            .style.format({"Cure %": "{:.1f}%"}),
            use_container_width=True, hide_index=True,
        )

    st.markdown("### 💚 Cured Accounts & Balance by Agent")
    cure_agent = cure_summary_by(filtered, "agent", "Agent").head(20)
    c1, c2 = st.columns(2)
    with c1:
        fig = px.bar(cure_agent.sort_values("Cured Accounts"), x="Cured Accounts", y="Agent",
                     orientation="h", title="Cured Accounts by Agent (Top 20)",
                     color="Cured Accounts", color_continuous_scale=GREENS_SCALE)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_23")
    with c2:
        fig = px.bar(cure_agent.sort_values("Cured Balance"), x="Cured Balance", y="Agent",
                     orientation="h", title="Cured Balance by Agent (Top 20)",
                     color="Cured Balance", color_continuous_scale=GREENS_SCALE)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_24")

    st.markdown("---")
    st.markdown("### 📅 PTP & BP by Agent")
    st.caption(
        "A record only counts as a PTP when **PTP Date** contains a valid value — blank/null "
        "PTP Dates are excluded from all PTP counts, amounts, and BP calculations below. "
        "PTP-to-Cured Conversion Rate = Cured Count ÷ PTP Count × 100. "
        "PTP-to-BP Conversion Rate = BP Count ÷ PTP Count × 100 (share of promises later broken; lower is better). "
        "RFD = agent's most frequent Reason for Default; RFD Value = OB across the agent's "
        "RFD-tagged accounts; PTP RFD / BP RFD = how many of those RFD-tagged accounts fall "
        "within the agent's PTP / BP subsets."
    )
    ptp_bp_agent = agent_ptp_bp_matrix(filtered)  # already computed above for KPIs; re-shown here for context
    st.dataframe(
        ptp_bp_agent.style.format({
            "Total Visits": "{:,.0f}", "PTP Amount": PHP, "BP OB": PHP, "RFD Value": PHP,
            "PTP-to-Cured Conversion Rate %": "{:.1f}%",
            "PTP-to-BP Conversion Rate %": lambda v: "—" if pd.isna(v) else f"{v:.1f}%",
        }),
        use_container_width=True, hide_index=True, height=380,
    )

    c1, c2 = st.columns(2)
    with c1:
        top_ptp = ptp_bp_agent.sort_values("PTP Amount", ascending=False).head(20)
        fig = px.bar(top_ptp.sort_values("PTP Amount"), x="PTP Amount", y="Agent",
                     orientation="h", title="Top 20 Agents by PTP Amount",
                     color="PTP Amount", color_continuous_scale=BLUES_SCALE)
        st.plotly_chart(fig, use_container_width=True, key="chart_agent_ptp_amount")
    with c2:
        top_bp = ptp_bp_agent.sort_values("BP Count", ascending=False).head(20)
        fig = px.bar(top_bp.sort_values("BP Count"), x="BP Count", y="Agent",
                     orientation="h", title="Top 20 Agents by BP Count",
                     color="BP Count", color_continuous_scale=REDS_SCALE)
        st.plotly_chart(fig, use_container_width=True, key="chart_agent_bp_count")

    st.markdown("### 📋 RFD (Reason for Default) Breakdown")
    rfd_all = rfd_summary_by(filtered, "all")
    rfd_ptp = rfd_summary_by(filtered, "ptp")
    rfd_bp = rfd_summary_by(filtered, "bp")
    if rfd_all.empty:
        st.info("No RFD (Reason for Default) values detected in this dataset.")
    else:
        rfd_tabs = st.tabs(["All Accounts", "PTP RFD", "BP RFD"])
        with rfd_tabs[0]:
            c1, c2 = st.columns([2, 3])
            with c1:
                st.dataframe(rfd_all.style.format({"Value": PHP}), use_container_width=True,
                             hide_index=True, height=340)
            with c2:
                fig = px.bar(rfd_all.sort_values("Count"), x="Count", y="RFD", orientation="h",
                             title="RFD Distribution — All Accounts", color="Count",
                             color_continuous_scale=BLUES_SCALE)
                st.plotly_chart(fig, use_container_width=True, key="chart_rfd_all")
        with rfd_tabs[1]:
            if rfd_ptp.empty:
                st.info("No RFD-tagged accounts within the PTP subset.")
            else:
                c1, c2 = st.columns([2, 3])
                with c1:
                    st.dataframe(rfd_ptp.style.format({"Value": PHP}), use_container_width=True,
                                 hide_index=True, height=340)
                with c2:
                    fig = px.bar(rfd_ptp.sort_values("Count"), x="Count", y="RFD", orientation="h",
                                 title="PTP RFD Distribution", color="Count",
                                 color_continuous_scale=BLUES_SCALE)
                    st.plotly_chart(fig, use_container_width=True, key="chart_rfd_ptp")
        with rfd_tabs[2]:
            if rfd_bp.empty:
                st.info("No RFD-tagged accounts within the BP subset.")
            else:
                c1, c2 = st.columns([2, 3])
                with c1:
                    st.dataframe(rfd_bp.style.format({"Value": PHP}), use_container_width=True,
                                 hide_index=True, height=340)
                with c2:
                    fig = px.bar(rfd_bp.sort_values("Count"), x="Count", y="RFD", orientation="h",
                                 title="BP RFD Distribution", color="Count",
                                 color_continuous_scale=REDS_SCALE)
                    st.plotly_chart(fig, use_container_width=True, key="chart_rfd_bp")

    with st.expander("🔎 Drill down: PTP records for a specific agent"):
        agent_options = sorted(ptp_bp_agent["Agent"].unique())
        if agent_options:
            sel_agent = st.selectbox("Select an Agent", agent_options, key="agent_ptp_drilldown")
            agent_ptp_rows = distinct_accounts(
                filtered[(filtered["agent"] == sel_agent) & (filtered["is_valid_ptp"])]
            )[["concat", "account_name", "action_date", "ptp_date", "ptp_amount", "ob_tad", "bp_status", "rfd"]].rename(
                columns={"concat": "Account Number", "account_name": "Account Name",
                         "action_date": "Action Date", "ptp_date": "PTP Date",
                         "ptp_amount": "PTP Amount", "ob_tad": "OB", "bp_status": "BP?", "rfd": "RFD"}
            ).sort_values("PTP Date", ascending=False)
            agent_ptp_rows["Action Date"] = agent_ptp_rows["Action Date"].apply(
                lambda d: d.strftime("%m/%d/%Y") if pd.notna(d) else "")
            agent_ptp_rows["PTP Date"] = agent_ptp_rows["PTP Date"].apply(
                lambda d: d.strftime("%m/%d/%Y") if pd.notna(d) else "")
            st.dataframe(agent_ptp_rows.style.format({"PTP Amount": PHP, "OB": PHP}),
                         use_container_width=True, hide_index=True, height=320)
        else:
            st.info("No agents with valid PTP records in the current filter selection.")

# ============================================================================
# TAB 4 — TELE & FIELD RESULTS DASHBOARD
# ============================================================================

with tabs[3]:
    st.subheader("Tele & Field Results Dashboard")

    tele_matrix = full_status_matrix(filtered, "substatus", "Tele Result")
    tele_body = tele_matrix[tele_matrix["Tele Result"] != "TOTAL"]
    field_matrix = full_status_matrix(filtered, "fv_result", "Field Result")
    field_body = field_matrix[field_matrix["Field Result"] != "TOTAL"]
    visits_acc = distinct_accounts(filtered)
    total_visits = visits_acc["total_visits"].sum()
    avg_visits = visits_acc["total_visits"].mean()

    st.markdown("#### 📌 Tele & Field KPIs")
    if not tele_body.empty and not field_body.empty:
        top_tele_row = tele_body.loc[tele_body["Total Count"].idxmax()]
        top_field_row = field_body.loc[field_body["Total Count"].idxmax()]
        k1, k2, k3, k4, k5 = st.columns(5)
        k1.metric("Tele Result Categories", fnum(tele_body["Tele Result"].nunique()))
        k2.metric("Field Result Categories", fnum(field_body["Field Result"].nunique()))
        k3.metric("Top Tele Result", top_tele_row["Tele Result"],
                   f"{fnum(top_tele_row['Total Count'])} accounts", delta_color="off")
        k4.metric("Top Field Result", top_field_row["Field Result"],
                   f"{fnum(top_field_row['Total Count'])} accounts", delta_color="off")
        k5.metric("Total Field Visits Made", fnum(total_visits),
                   f"Avg {avg_visits:.2f}/account" if pd.notna(avg_visits) else "Avg 0.00/account",
                   delta_color="off")
    else:
        st.info("Not enough Tele/Field Result data in the current filter selection to compute KPIs.")

    st.markdown("### 📞 Tele Result (SubStatus)")
    money_cols = [c for c in tele_matrix.columns if "Balance" in c]
    st.dataframe(
        tele_matrix.style.format({c: PHP for c in money_cols}),
        use_container_width=True, hide_index=True, height=380,
    )

    st.markdown("### 🚶 Field Result (Fv_Result)")
    money_cols = [c for c in field_matrix.columns if "Balance" in c]
    st.dataframe(
        field_matrix.style.format({c: PHP for c in money_cols}),
        use_container_width=True, hide_index=True, height=460,
    )

    # ---- Total of Visits Made ----
    st.markdown("### 🚗 Field Visit Volume (Total of Visits Made)")
    r1 = st.columns(2)
    r1[0].metric("Total Visits", fnum(total_visits))
    r1[1].metric("Average Visits per Account", f"{avg_visits:.2f}" if pd.notna(avg_visits) else "0.00")

    visits_per_result = filtered.dropna(subset=["fv_result"])
    visits_per_result = visits_per_result[visits_per_result["fv_result"] != "Unspecified"]
    vpr = distinct_accounts(visits_per_result).groupby("fv_result").agg(
        Total_Visits=("total_visits", "sum"), Avg_Visits=("total_visits", "mean"),
        Accounts=("concat", "nunique"),
    ).reset_index().rename(columns={"fv_result": "Field Result", "Total_Visits": "Total Visits",
                                     "Avg_Visits": "Average Visits"})
    vpr = vpr.sort_values("Total Visits", ascending=False)
    st.dataframe(vpr.style.format({"Average Visits": "{:.2f}"}), use_container_width=True, hide_index=True)

    st.markdown("### 📊 Visualizations")
    c1, c2 = st.columns(2)
    with c1:
        fig = px.pie(tele_body, names="Tele Result", values="Total Count", hole=0.4,
                     title="Tele Result Distribution", color_discrete_sequence=BLUES_SCALE)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_25")
    with c2:
        fig = px.pie(field_body.sort_values("Total Count", ascending=False).head(12),
                     names="Field Result", values="Total Count", hole=0.4,
                     title="Field Result Distribution (Top 12)", color_discrete_sequence=GREENS_SCALE)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_26")

    c1, c2 = st.columns(2)
    with c1:
        m = tele_body[["Tele Result", "Active Count", "Cured Count", "Flowed Count"]].melt(
            id_vars="Tele Result", value_vars=["Active Count", "Cured Count", "Flowed Count"],
            var_name="Status", value_name="Accounts",
        )
        m["Status"] = m["Status"].str.replace(" Count", "")
        fig = px.bar(m.sort_values("Accounts"), x="Accounts", y="Tele Result", color="Status",
                     orientation="h", barmode="stack", color_discrete_map=STATUS_COLORS,
                     title="Tele Result vs Status")
        fig.update_layout(height=450)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_27")
    with c2:
        top_field = field_body.sort_values("Total Count", ascending=False).head(15)["Field Result"]
        m = field_body[field_body["Field Result"].isin(top_field)][
            ["Field Result", "Active Count", "Cured Count", "Flowed Count"]].melt(
            id_vars="Field Result", value_vars=["Active Count", "Cured Count", "Flowed Count"],
            var_name="Status", value_name="Accounts",
        )
        m["Status"] = m["Status"].str.replace(" Count", "")
        fig = px.bar(m.sort_values("Accounts"), x="Accounts", y="Field Result", color="Status",
                     orientation="h", barmode="stack", color_discrete_map=STATUS_COLORS,
                     title="Field Result vs Status (Top 15)")
        fig.update_layout(height=450)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_28")

    c1, c2 = st.columns(2)
    with c1:
        fig = px.bar(vpr.head(15), x="Field Result", y="Average Visits",
                     title="Total Visit Analysis — Average Visits per Field Result",
                     color="Average Visits", color_continuous_scale=BLUES_SCALE)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_29")
    with c2:
        top_ob = pd.concat([
            tele_body.rename(columns={"Tele Result": "Result"})[["Result", "Total Balance"]].assign(Type="Tele"),
            field_body.rename(columns={"Field Result": "Result"})[["Result", "Total Balance"]].assign(Type="Field"),
        ]).sort_values("Total Balance", ascending=False).head(15)
        fig = px.bar(top_ob.sort_values("Total Balance"), x="Total Balance", y="Result", color="Type",
                     orientation="h", title="Top Results by OB (Tele + Field)",
                     color_discrete_map={"Tele": SB_BLUE, "Field": SB_GREEN})
        st.plotly_chart(fig, use_container_width=True, key=f"chart_30")

# ============================================================================
# TAB 5 — AREA DASHBOARD
# ============================================================================

with tabs[4]:
    st.subheader("Area Performance")
    st.caption("Area data is sourced from the **AREA2** field.")
    area_mat = full_status_matrix(filtered, "area2", "AREA2")
    area_mat["Cure Rate %"] = np.where(
        area_mat["Total Count"] > 0, area_mat["Cured Count"] / area_mat["Total Count"] * 100, 0
    )
    area_body = area_mat[area_mat["AREA2"] != "TOTAL"]

    st.markdown("#### 📌 Area KPIs")
    if not area_body.empty:
        top_bal_row = area_body.loc[area_body["Total Balance"].idxmax()]
        top_cure_row = area_body.loc[area_body["Cure Rate %"].idxmax()]
        low_cure_row = area_body.loc[area_body["Cure Rate %"].idxmin()]
        avg_cure_rate = area_body["Cure Rate %"].mean()

        k1, k2, k3, k4, k5 = st.columns(5)
        k1.metric("Total Areas", fnum(area_body["AREA2"].nunique()))
        k2.metric("Top Area by Balance", top_bal_row["AREA2"], PHP(top_bal_row["Total Balance"]))
        k3.metric("Best Cure Rate", top_cure_row["AREA2"], pct(top_cure_row["Cure Rate %"]))
        k4.metric("⚠️ Needs Attention (Lowest Cure Rate)", low_cure_row["AREA2"],
                   pct(low_cure_row["Cure Rate %"]), delta_color="inverse")
        k5.metric("Avg Cure Rate Across Areas", pct(avg_cure_rate))
    else:
        st.info("Not enough Area data in the current filter selection to compute KPIs.")

    st.markdown("#### 📊 Visit & Conversion KPIs")
    area_acc = distinct_accounts(filtered)
    total_accounts_area = area_acc["concat"].nunique()
    total_visits_area = area_acc["total_visits"].sum()
    unique_total_visits_area = area_acc[area_acc["total_visits"] > 0]["concat"].nunique()
    ptp_acc_area = area_acc[area_acc["is_valid_ptp"]]
    ptp_count_area = ptp_acc_area["concat"].nunique()
    cured_acc_area = area_acc[area_acc["status_norm"] == "Cured"]
    cure_count_area = cured_acc_area["concat"].nunique()
    cure_rate_area = (cure_count_area / total_accounts_area * 100) if total_accounts_area else 0
    conversion_rate_area = (cure_count_area / ptp_count_area * 100) if ptp_count_area else 0

    j1, j2, j3, j4, j5, j6 = st.columns(6)
    j1.metric("Total Visits", fnum(total_visits_area))
    j2.metric("Unique Total Visits", fnum(unique_total_visits_area))
    j3.metric("PTP Count", fnum(ptp_count_area))
    j4.metric("Cure Count", fnum(cure_count_area))
    render_rate_metric(j5, "Cure Rate", pct(cure_rate_area), cure_rate_area, good=70, warn=40)
    render_rate_metric(j6, "Conversion Rate (PTP → Cured)", pct(conversion_rate_area),
                        conversion_rate_area, good=70, warn=40)
    st.caption(
        "Total Visits sums the **Total of Visits Made** field across all accounts. Unique Total "
        "Visits counts distinct accounts with at least one recorded visit. "
        "Conversion Rate = Cure Count ÷ PTP Count × 100."
    )

    money_cols = [c for c in area_mat.columns if "Balance" in c]
    st.dataframe(
        area_mat.style.format({**{c: PHP for c in money_cols}, "Cure Rate %": "{:.1f}%"}),
        use_container_width=True, hide_index=True, height=380,
    )
    c1, c2 = st.columns(2)
    with c1:
        fig = px.bar(area_body.sort_values("Total Balance"), x="Total Balance", y="AREA2",
                     orientation="h", title="Area Ranking by Balance", color="Total Balance",
                     color_continuous_scale=BLUES_SCALE)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_31")
    with c2:
        m = area_body[["AREA2", "Active Count", "Cured Count", "Flowed Count"]].melt(
            id_vars="AREA2", value_vars=["Active Count", "Cured Count", "Flowed Count"],
            var_name="Status", value_name="Accounts",
        )
        m["Status"] = m["Status"].str.replace(" Count", "")
        fig = px.bar(m, x="AREA2", y="Accounts", color="Status", barmode="stack",
                     color_discrete_map=STATUS_COLORS, title="Area Status Breakdown")
        st.plotly_chart(fig, use_container_width=True, key=f"chart_32")

    fig = px.treemap(area_body, path=["AREA2"], values="Total Balance",
                      color="Total Balance", color_continuous_scale=BLUES_SCALE,
                      title="Area Treemap (by Balance)")
    st.plotly_chart(fig, use_container_width=True, key=f"chart_33")

    st.markdown("### 🔵🟢 Area Break — NCR vs Universal")
    st.caption("Area Break grouping is sourced from the **Area List** column (falling back to Area, then Area2, if Area List has no NCR/Universal matches).")
    a_list = filtered.copy()
    a_list["area_group"] = a_list["area_list"].apply(
        lambda v: "NCR" if "ncr" in str(v).lower()
        else ("Universal" if "univ" in str(v).lower() else None)
    )
    a2 = filtered.copy()
    a2["area_group"] = a2["area"].apply(
        lambda v: "NCR" if "ncr" in str(v).lower()
        else ("Universal" if "univ" in str(v).lower() else None)
    )
    a2_fallback = filtered.copy()
    a2_fallback["area_group"] = a2_fallback["area2"].apply(
        lambda v: "NCR" if "ncr" in str(v).lower()
        else ("Universal" if "univ" in str(v).lower() else None)
    )
    grp = a_list[a_list["area_group"].notna()]
    if grp.empty:
        grp = a2[a2["area_group"].notna()]
    if grp.empty:
        grp = a2_fallback[a2_fallback["area_group"].notna()]

    if not grp.empty:
        rows = []
        for name, g in grp.groupby("area_group"):
            acc_g = distinct_accounts(g)
            row = {"Area Break": name, "Total Accounts": acc_g["concat"].nunique(),
                   "Total Balance": acc_g["ob_tad"].sum()}
            for st_name in ["Active", "Flowed", "Cured"]:
                s = acc_g[acc_g["status_norm"] == st_name]
                row[f"{st_name} Accounts"] = s["concat"].nunique()
                row[f"{st_name} Balance"] = s["ob_tad"].sum()
            rows.append(row)
        comp = pd.DataFrame(rows)
        money_cols = [c for c in comp.columns if "Balance" in c]
        st.dataframe(comp.style.format({c: PHP for c in money_cols}),
                     use_container_width=True, hide_index=True)
        c1, c2 = st.columns(2)
        with c1:
            m = comp[["Area Break", "Active Accounts", "Flowed Accounts", "Cured Accounts"]].melt(
                id_vars="Area Break",
                value_vars=["Active Accounts", "Flowed Accounts", "Cured Accounts"],
                var_name="Status", value_name="Accounts",
            )
            m["Status"] = m["Status"].str.replace(" Accounts", "")
            fig = px.bar(m, x="Area Break", y="Accounts", color="Status", barmode="group",
                         color_discrete_map=STATUS_COLORS, title="NCR vs Universal — Accounts")
            st.plotly_chart(fig, use_container_width=True, key=f"chart_34")
        with c2:
            m = comp[["Area Break", "Active Balance", "Flowed Balance", "Cured Balance"]].melt(
                id_vars="Area Break",
                value_vars=["Active Balance", "Flowed Balance", "Cured Balance"],
                var_name="Status", value_name="Balance",
            )
            m["Status"] = m["Status"].str.replace(" Balance", "")
            fig = px.bar(m, x="Area Break", y="Balance", color="Status", barmode="group",
                         color_discrete_map=STATUS_COLORS, title="NCR vs Universal — Balance")
            st.plotly_chart(fig, use_container_width=True, key=f"chart_35")
    else:
        st.info("No records found tagged as NCR or Universal in the Area List, Area, or Area2 columns.")

    st.markdown("### Provincial / Sub-Area Breakdown")
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**Active Areas**")
        prov_active = summarize_by(filtered[filtered["status_norm"] == "Active"], "sub_area")
        prov_active = prov_active.rename(columns={"sub_area": "Province"})[["Province", "Accounts", "Balance"]]
        total_row = pd.DataFrame([{"Province": "TOTAL", "Accounts": prov_active["Accounts"].sum(),
                                    "Balance": prov_active["Balance"].sum()}])
        st.dataframe(pd.concat([prov_active, total_row], ignore_index=True)
                     .style.format({"Balance": PHP}), use_container_width=True, hide_index=True, height=350)
    with c2:
        st.markdown("**Flowed Areas**")
        prov_flowed = summarize_by(filtered[filtered["status_norm"] == "Flowed"], "sub_area")
        prov_flowed = prov_flowed.rename(columns={"sub_area": "Province"})[["Province", "Accounts", "Balance"]]
        total_row = pd.DataFrame([{"Province": "TOTAL", "Accounts": prov_flowed["Accounts"].sum(),
                                    "Balance": prov_flowed["Balance"].sum()}])
        st.dataframe(pd.concat([prov_flowed, total_row], ignore_index=True)
                     .style.format({"Balance": PHP}), use_container_width=True, hide_index=True, height=350)

    st.markdown("### 💚 Cured Accounts & Balance by Area")
    cure_area = cure_summary_by(filtered, "area2", "AREA2")
    c1, c2 = st.columns(2)
    with c1:
        fig = px.bar(cure_area.sort_values("Cured Accounts"), x="Cured Accounts", y="AREA2",
                     orientation="h", title="Cured Accounts by Area", color="Cured Accounts",
                     color_continuous_scale=GREENS_SCALE)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_36")
    with c2:
        fig = px.bar(cure_area.sort_values("Cured Balance"), x="Cured Balance", y="AREA2",
                     orientation="h", title="Cured Balance by Area", color="Cured Balance",
                     color_continuous_scale=GREENS_SCALE)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_37")

# ============================================================================
# TAB 6 — AREA MAP DASHBOARD (Area List)
# ============================================================================

with tabs[5]:
    st.subheader("Area Map Dashboard")
    st.caption("Geographic view based on the **Area List** column, aggregated to Philippine province/region centroids.")

    map_acc = distinct_accounts(filtered)
    total_accounts_map = map_acc["concat"].nunique()
    total_visits_map = map_acc["total_visits"].sum()
    avg_visits_map = map_acc["total_visits"].mean()
    visited_acc = map_acc[map_acc["total_visits"] > 0]
    visited_count = visited_acc["concat"].nunique()
    coverage_rate = (visited_count / total_accounts_map * 100) if total_accounts_map else 0
    ptp_acc_map = map_acc[map_acc["is_valid_ptp"]]
    ptp_count_map = ptp_acc_map["concat"].nunique()
    cured_acc_map = map_acc[map_acc["status_norm"] == "Cured"]
    cure_count_map = cured_acc_map["concat"].nunique()
    conversion_rate_map = (cure_count_map / ptp_count_map * 100) if ptp_count_map else 0

    st.markdown("#### 📌 Visit & Coverage KPIs")
    j1, j2, j3, j4 = st.columns(4)
    j1.metric("Total Visit Count", fnum(total_visits_map))
    j2.metric("Average Visit Count", f"{avg_visits_map:,.2f}" if pd.notna(avg_visits_map) else "0.00")
    j3.metric("Visited Accounts", fnum(visited_count))
    j4.metric("PTP Count", fnum(ptp_count_map))
    j5, j6 = st.columns(2)
    j5.metric("Cure Count", fnum(cure_count_map))
    render_rate_metric(j6, "PTP-to-Cure Conversion Rate", pct(conversion_rate_map),
                        conversion_rate_map, good=70, warn=40)
    k1, _ = st.columns(2)
    render_rate_metric(k1, "Coverage Rate (Visited ÷ Total Accounts)", pct(coverage_rate),
                        coverage_rate, good=80, warn=50)
    st.markdown("---")

    map_summary = full_status_matrix(filtered, "area_list", "Area List")
    map_body = map_summary[map_summary["Area List"] != "TOTAL"].copy()

    if map_body.empty:
        st.info("No usable Area List values found in the current filter selection.")
    else:
        coords = map_body["Area List"].apply(geocode_area)
        map_body["lat"] = coords.apply(lambda c: c[0] if c else np.nan)
        map_body["lon"] = coords.apply(lambda c: c[1] if c else np.nan)
        map_body["Cure Rate %"] = np.where(
            map_body["Total Count"] > 0, map_body["Cured Count"] / map_body["Total Count"] * 100, 0
        )

        mapped = map_body.dropna(subset=["lat", "lon"])
        unmapped = map_body[map_body["lat"].isna()]

        r1 = st.columns(3)
        r1[0].metric("Area List Values", fnum(len(map_body)))
        r1[1].metric("Mapped to Coordinates", fnum(len(mapped)))
        r1[2].metric("Unmapped Labels", fnum(len(unmapped)))

        if not mapped.empty:
            metric_choice = st.radio(
                "Bubble size represents:", ["Total Accounts", "Total Balance", "Cure Rate %"],
                horizontal=True, key="map_metric_choice",
            )
            size_col = {"Total Accounts": "Total Count", "Total Balance": "Total Balance",
                        "Cure Rate %": "Cure Rate %"}[metric_choice]

            scatter_fn = px.scatter_map if hasattr(px, "scatter_map") else px.scatter_mapbox
            map_style_kwarg = "map_style" if hasattr(px, "scatter_map") else "mapbox_style"
            fig = scatter_fn(
                mapped, lat="lat", lon="lon", size=size_col, color="Cure Rate %",
                color_continuous_scale=GREENS_SCALE, size_max=45, zoom=4.4,
                hover_name="Area List",
                hover_data={"Total Count": True, "Total Balance": ":,.2f", "Cured Count": True,
                            "Active Count": True, "Flowed Count": True, "Cure Rate %": ":.1f",
                            "lat": False, "lon": False},
                title="Area List — Portfolio Map (Philippines)",
            )
            fig.update_layout(**{map_style_kwarg: "open-street-map"}, height=560,
                               margin={"r": 0, "t": 40, "l": 0, "b": 0})
            st.plotly_chart(fig, use_container_width=True, key="chart_map_1")
        else:
            st.info("None of the Area List values could be matched to a known Philippine province/region.")

        st.markdown("### Area List — Full Matrix")
        money_cols = [c for c in map_body.columns if "Balance" in c]
        st.dataframe(
            map_body.drop(columns=["lat", "lon"]).style.format(
                {**{c: PHP for c in money_cols}, "Cure Rate %": "{:.1f}%"}),
            use_container_width=True, hide_index=True, height=380,
        )

        if not unmapped.empty:
            with st.expander(f"⚠️ {len(unmapped)} Area List label(s) could not be matched to a map location"):
                st.dataframe(unmapped[["Area List", "Total Count", "Total Balance"]]
                             .style.format({"Total Balance": PHP}), use_container_width=True, hide_index=True)
                st.caption("These are still included in every other dashboard/table — they're only excluded from the map bubbles above.")

# ============================================================================
# TAB 7 — ACTIVE STATUS DASHBOARD (Status = Active only)
# ============================================================================

with tabs[6]:
    st.subheader("Active Status Dashboard")
    st.caption("This dashboard is scoped to **Status = Active** accounts only, regardless of the Status filter above.")

    active_only = filtered[filtered["status_norm"] == "Active"]

    if active_only.empty:
        st.info("No Active-status records in the current filter selection.")
    else:
        active_acc_ct = distinct_accounts(active_only)["concat"].nunique()
        active_bal_ct = distinct_accounts(active_only)["ob_tad"].sum()
        r1 = st.columns(2)
        r1[0].metric("Active Accounts", fnum(active_acc_ct))
        r1[1].metric("Active Balance", PHP(active_bal_ct))

        st.markdown("### A. Active Status (from `active` column)")
        act_sum = summarize_by(active_only, "active").rename(columns={"active": "Active Status"})
        c1, c2 = st.columns(2)
        with c1:
            fig = px.pie(act_sum, names="Active Status", values="Accounts", hole=0.5,
                         title="Active Status — Donut", color_discrete_sequence=BLUES_SCALE)
            st.plotly_chart(fig, use_container_width=True, key=f"chart_38")
        with c2:
            fig = px.bar(act_sum.sort_values("Accounts"), x="Accounts", y="Active Status",
                         orientation="h", title="Active Status Breakdown", color="Accounts",
                         color_continuous_scale=BLUES_SCALE)
            st.plotly_chart(fig, use_container_width=True, key=f"chart_39")
        st.dataframe(act_sum[["Active Status", "Accounts", "Balance"]].rename(
            columns={"Accounts": "Count"}).style.format({"Balance": PHP}),
            use_container_width=True, hide_index=True)

        st.markdown("### B. Tele Result (from SubStatus)")
        tele_sum = summarize_by(active_only, "substatus").rename(columns={"substatus": "Tele Result"})
        c1, c2 = st.columns(2)
        with c1:
            fig = px.bar(tele_sum.sort_values("Accounts"), x="Accounts", y="Tele Result",
                         orientation="h", title="Tele Result Breakdown (Active)", color="Accounts",
                         color_continuous_scale=BLUES_SCALE)
            st.plotly_chart(fig, use_container_width=True, key=f"chart_40")
        with c2:
            st.dataframe(tele_sum[["Tele Result", "Accounts", "Balance"]].rename(
                columns={"Accounts": "Count"}).style.format({"Balance": PHP}),
                use_container_width=True, hide_index=True, height=380)

        st.markdown("### C. Field Result (from Fv_Result)")
        field_sum = summarize_by(active_only, "fv_result").rename(columns={"fv_result": "Field Result"})
        c1, c2 = st.columns(2)
        with c1:
            fig = px.bar(field_sum.sort_values("Accounts"), x="Accounts", y="Field Result",
                         orientation="h", title="Field Result Breakdown (Active)", color="Accounts",
                         color_continuous_scale=GREENS_SCALE)
            st.plotly_chart(fig, use_container_width=True, key=f"chart_41")
        with c2:
            st.dataframe(field_sum[["Field Result", "Accounts", "Balance"]].rename(
                columns={"Accounts": "Count"}).style.format({"Balance": PHP}),
                use_container_width=True, hide_index=True, height=380)

        st.markdown("### 🔥 Tele Result × Field Result Heatmap")
        heat_df = active_only.dropna(subset=["substatus", "fv_result"])
        heat_df = heat_df[(heat_df["substatus"] != "Unspecified") & (heat_df["fv_result"] != "Unspecified")]
        if not heat_df.empty:
            heat_acc = distinct_accounts(heat_df)
            pivot = heat_acc.pivot_table(index="substatus", columns="fv_result", values="concat",
                                          aggfunc="nunique", fill_value=0)
            top_rows = pivot.sum(axis=1).sort_values(ascending=False).head(12).index
            top_cols = pivot.sum(axis=0).sort_values(ascending=False).head(12).index
            pivot = pivot.loc[top_rows, top_cols]
            fig = px.imshow(pivot, text_auto=True, color_continuous_scale=BLUES_SCALE,
                             title="Tele Result × Field Result Heatmap (Active, Top Categories)",
                             aspect="auto", labels=dict(x="Field Result", y="Tele Result", color="Accounts"))
            st.plotly_chart(fig, use_container_width=True, key=f"chart_42")
        else:
            st.info("Not enough Tele Result / Field Result data to build the heatmap.")

# ============================================================================
# TAB 8 — CURED / PAYMENT DASHBOARD
# ============================================================================

with tabs[7]:
    st.subheader("Cured / Payment Dashboard")
    st.caption("Accounts with **Status = CURED** are treated as paid/cured accounts throughout this section.")

    acc = distinct_accounts(filtered)
    total_accounts = acc["concat"].nunique()
    total_balance = acc["ob_tad"].sum()
    cured_acc_df = acc[acc["status_norm"] == "Cured"]
    cured_accounts = cured_acc_df["concat"].nunique()
    cured_balance = cured_acc_df["ob_tad"].sum()
    cure_rate = (cured_accounts / total_accounts * 100) if total_accounts else 0
    cure_bal_rate = (cured_balance / total_balance * 100) if total_balance else 0

    r1 = st.columns(4)
    r1[0].metric("Total Cured Accounts", fnum(cured_accounts))
    r1[1].metric("Total Cured Balance", PHP(cured_balance))
    r1[2].metric("Cure Rate", pct(cure_rate))
    r1[3].metric("Cure Balance Rate", pct(cure_bal_rate))

    st.markdown("### Cured Analysis by Industry")
    cure_ind = cure_summary_by(filtered, "industry", "Industry")
    st.dataframe(
        cure_ind.style.format({"Cured Balance": PHP, "Cure Rate %": "{:.1f}%"}),
        use_container_width=True, hide_index=True,
    )

    st.markdown("### Cured Analysis by Agent")
    cure_agent = cure_summary_by(filtered, "agent", "Agent")
    st.dataframe(
        cure_agent.style.format({"Cured Balance": PHP, "Cure Rate %": "{:.1f}%"}),
        use_container_width=True, hide_index=True, height=350,
    )

    st.markdown("### Cured Analysis by Area")
    st.caption("Area data is sourced from the **AREA2** field.")
    cure_area = cure_summary_by(filtered, "area2", "AREA2")
    st.dataframe(
        cure_area.style.format({"Cured Balance": PHP, "Cure Rate %": "{:.1f}%"}),
        use_container_width=True, hide_index=True,
    )

    st.markdown("### 💚 Cured Accounts & Balance — Consolidated Charts")
    c1, c2, c3 = st.columns(3)
    with c1:
        fig = px.bar(cure_ind.sort_values("Cured Accounts"), x="Cured Accounts", y="Industry",
                     orientation="h", title="Cured Accounts by Industry", color="Cured Accounts",
                     color_continuous_scale=GREENS_SCALE)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_43")
    with c2:
        fig = px.bar(cure_agent.sort_values("Cured Accounts").tail(15), x="Cured Accounts", y="Agent",
                     orientation="h", title="Cured Accounts by Agent (Top 15)", color="Cured Accounts",
                     color_continuous_scale=GREENS_SCALE)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_44")
    with c3:
        fig = px.bar(cure_area.sort_values("Cured Accounts"), x="Cured Accounts", y="AREA2",
                     orientation="h", title="Cured Accounts by Area", color="Cured Accounts",
                     color_continuous_scale=GREENS_SCALE)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_45")

    c4, c5, c6 = st.columns(3)
    with c4:
        fig = px.bar(cure_ind.sort_values("Cured Balance"), x="Cured Balance", y="Industry",
                     orientation="h", title="Cured Balance by Industry", color="Cured Balance",
                     color_continuous_scale=GREENS_SCALE)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_46")
    with c5:
        fig = px.bar(cure_agent.sort_values("Cured Balance").tail(15), x="Cured Balance", y="Agent",
                     orientation="h", title="Cured Balance by Agent (Top 15)", color="Cured Balance",
                     color_continuous_scale=GREENS_SCALE)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_47")
    with c6:
        fig = px.bar(cure_area.sort_values("Cured Balance"), x="Cured Balance", y="AREA2",
                     orientation="h", title="Cured Balance by Area", color="Cured Balance",
                     color_continuous_scale=GREENS_SCALE)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_48")

    if filtered["action_date"].notna().any():
        st.markdown("### Cure Trend Over Time")
        trend = filtered.dropna(subset=["action_date"]).copy()
        trend["Month"] = trend["action_date"].dt.to_period("M").astype(str)
        cure_trend = trend[trend["status_norm"] == "Cured"].drop_duplicates(subset=["concat", "Month"]) \
            .groupby("Month").agg(Accounts=("concat", "nunique"), Balance=("ob_tad", "sum")).reset_index()
        c1, c2 = st.columns(2)
        with c1:
            fig = px.line(cure_trend, x="Month", y="Accounts", markers=True,
                          title="Cured Accounts Trend (Monthly)", color_discrete_sequence=[SB_GREEN])
            st.plotly_chart(fig, use_container_width=True, key=f"chart_49")
        with c2:
            fig = px.line(cure_trend, x="Month", y="Balance", markers=True,
                          title="Cured Balance Trend (Monthly)", color_discrete_sequence=[SB_GREEN])
            st.plotly_chart(fig, use_container_width=True, key=f"chart_50")

# ============================================================================
# TAB 9 — ACCOUNT DETAILS
# ============================================================================

with tabs[8]:
    st.subheader("Account Detail View")
    detail_cols = {
        "concat": "Account Number", "account_name": "Account Name", "industry": "Industry",
        "agent": "Agent", "ob_tad": "OB", "status": "Status", "substatus": "Tele Results",
        "fv_result": "FV Result", "area": "Area", "area_list": "Area Break",
        "endorsement_date": "Date Endorsed", "action_date": "Action Date",
        "ptp_date": "PTP Date", "ptp_amount": "PTP Amount", "bp_status": "BP?",
    }
    detail = distinct_accounts(filtered)[list(detail_cols.keys())].rename(columns=detail_cols)
    detail["Date Endorsed"] = detail["Date Endorsed"].apply(
        lambda d: d.strftime("%m/%d/%Y") if pd.notna(d) else ""
    )
    detail["Action Date"] = detail["Action Date"].apply(
        lambda d: d.strftime("%m/%d/%Y") if pd.notna(d) else ""
    )
    detail["PTP Date"] = detail["PTP Date"].apply(
        lambda d: d.strftime("%m/%d/%Y") if pd.notna(d) else ""
    )
    st.caption(f"{len(detail):,} distinct accounts")

    if HAS_AGGRID:
        gb = GridOptionsBuilder.from_dataframe(detail)
        gb.configure_default_column(filterable=True, sortable=True, resizable=True)
        gb.configure_pagination(paginationAutoPageSize=False, paginationPageSize=25)
        gb.configure_side_bar()
        AgGrid(detail, gridOptions=gb.build(), update_mode=GridUpdateMode.NO_UPDATE,
               theme="balham", height=480, fit_columns_on_grid_load=True)
    else:
        st.dataframe(detail, use_container_width=True, height=480)
        st.caption("Install `streamlit-aggrid` for advanced search/sort/filter grid features.")

    st.download_button(
        "⬇️ Export Account Details (Excel)",
        data=to_excel_bytes({"Account Details": detail}),
        file_name="account_details.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ============================================================================
# TAB 10 — DOWNLOADS
# ============================================================================

with tabs[9]:
    st.subheader("Export Center")
    st.caption("All exports reflect the currently applied filters and include Security Bank-themed styling.")

    c1, c2 = st.columns(2)
    with c1:
        summary_sheets = {
            "Executive Summary": status_summary(filtered),
            "Industry": full_status_matrix(filtered, "industry", "INDUSTRY"),
            "Agent": agent_matrix(filtered),
        }
        st.download_button("⬇️ Download Dashboard Summary", to_excel_bytes(summary_sheets),
                            "dashboard_summary.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        st.download_button("⬇️ Download Industry Report",
                            to_excel_bytes({"Industry": full_status_matrix(filtered, "industry", "INDUSTRY")}),
                            "industry_report.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        area_sheets = {
            "Area (AREA2)": full_status_matrix(filtered, "area2", "AREA2"),
            "Area Break (Area List)": summarize_by(filtered, "area_list"),
            "Sub Area": summarize_by(filtered, "sub_area"),
        }
        st.download_button("⬇️ Download Area Report", to_excel_bytes(area_sheets),
                            "area_report.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with c2:
        st.download_button("⬇️ Download Agent Performance",
                            to_excel_bytes({"Agent": agent_matrix(filtered)}),
                            "agent_performance.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        tele_field_sheets = {
            "Tele Result": full_status_matrix(filtered, "substatus", "Tele Result"),
            "Field Result": full_status_matrix(filtered, "fv_result", "Field Result"),
        }
        st.download_button("⬇️ Download Tele & Field Results Report", to_excel_bytes(tele_field_sheets),
                            "tele_field_results_report.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        cured_sheets = {
            "Cured by Industry": cure_summary_by(filtered, "industry", "Industry"),
            "Cured by Agent": cure_summary_by(filtered, "agent", "Agent"),
            "Cured by Area (AREA2)": cure_summary_by(filtered, "area2", "AREA2"),
        }
        st.download_button("⬇️ Download Cured/Payment Report", to_excel_bytes(cured_sheets),
                            "cured_payment_report.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        exec_sheets = {
            "BOM-FRESH": summarize_by(filtered, "bom_fresh"),
            "Balance Distribution": full_status_matrix(filtered, "bal_distro", "BAL Distro"),
        }
        st.download_button("⬇️ Download Executive (BOM/FRESH & Balance Distro) Report",
                            to_excel_bytes(exec_sheets), "executive_bom_fresh_baldistro.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        bp_export_sheets = {
            "Agent PTP & BP": agent_ptp_bp_matrix(filtered),
        }
        st.download_button("⬇️ Download Agent PTP & BP Report", to_excel_bytes(bp_export_sheets),
                            "agent_ptp_bp_report.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        st.download_button("⬇️ Download Raw (Filtered) Data",
                            to_excel_bytes({"Raw Data": filtered.drop(columns=["status_norm", "has_payment"])}),
                            "raw_data_filtered.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

st.markdown("---")
st.caption("SBC B2 BEL Collection Dashboard · Security Bank · Banking-grade curing & portfolio performance analytics")
